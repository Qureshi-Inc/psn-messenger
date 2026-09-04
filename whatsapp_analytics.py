"""
WhatsApp message storage and analytics engine.

Parsing approach adapted from JBoixCampos/whatsapp-chat-analyzer
(https://github.com/JBoixCampos/whatsapp-chat-analyzer, MIT License).
Copyright (c) 2025 Javier Boix Campos — MIT License.
"""

from __future__ import annotations

import calendar
import hashlib
import io
import logging
import os
import re
import sqlite3
import threading
import time
import zipfile
from collections import Counter, defaultdict
from datetime import datetime
from pathlib import Path
from typing import Optional
from zoneinfo import ZoneInfo

logger = logging.getLogger(__name__)

_DB_PATH = Path("/data/whatsapp.db")
_lock = threading.Lock()
_TZ = ZoneInfo("America/Los_Angeles")


def _load_aliases() -> dict[str, str]:
    """Parse WA_NAME_ALIASES env var.

    Format: comma-separated ``old:new`` pairs, e.g.
    ``@interestingsoup:Moiz,oldname:NewName``
    Matching is case-insensitive; stored names are replaced with the canonical value.
    """
    raw = os.environ.get("WA_NAME_ALIASES", "").strip()
    result: dict[str, str] = {}
    if not raw:
        return result
    for pair in raw.split(","):
        pair = pair.strip()
        if ":" not in pair:
            continue
        old, _, new = pair.partition(":")
        old, new = old.strip(), new.strip()
        if old and new:
            result[old.lower()] = new
    return result


_NAME_ALIASES: dict[str, str] = {}


def _resolve_name(name: str) -> str:
    return _NAME_ALIASES.get(name.lower(), name)

MAX_UPLOAD_BYTES = 50 * 1024 * 1024
MAX_ZIP_UNCOMPRESSED = 150 * 1024 * 1024
MAX_ZIP_FILES = 10


# ── Database ───────────────────────────────────────────────────────────────────

def _conn() -> sqlite3.Connection:
    c = sqlite3.connect(_DB_PATH, check_same_thread=False)
    c.row_factory = sqlite3.Row
    return c


def init() -> None:
    global _NAME_ALIASES
    _NAME_ALIASES = _load_aliases()
    _DB_PATH.parent.mkdir(parents=True, exist_ok=True)
    with _lock, _conn() as db:
        db.executescript("""
        CREATE TABLE IF NOT EXISTS whatsapp_messages (
            id               TEXT PRIMARY KEY,
            message_id       TEXT,
            group_jid        TEXT,
            sender_jid       TEXT,
            sender_name      TEXT NOT NULL,
            timestamp        INTEGER NOT NULL,
            text             TEXT,
            message_type     TEXT NOT NULL DEFAULT 'text',
            has_photo        INTEGER NOT NULL DEFAULT 0,
            has_video        INTEGER NOT NULL DEFAULT 0,
            has_audio        INTEGER NOT NULL DEFAULT 0,
            has_document     INTEGER NOT NULL DEFAULT 0,
            is_media_omitted INTEGER NOT NULL DEFAULT 0,
            reply_to         TEXT,
            from_me          INTEGER NOT NULL DEFAULT 0,
            source           TEXT NOT NULL DEFAULT 'historical_export',
            import_batch     TEXT,
            created_at       INTEGER NOT NULL
        );
        CREATE INDEX IF NOT EXISTS idx_wa_ts     ON whatsapp_messages(timestamp);
        CREATE INDEX IF NOT EXISTS idx_wa_sender ON whatsapp_messages(sender_name, timestamp);
        CREATE INDEX IF NOT EXISTS idx_wa_msgid  ON whatsapp_messages(message_id);

        CREATE TABLE IF NOT EXISTS whatsapp_reactions (
            id           TEXT PRIMARY KEY,
            target_msg_id TEXT NOT NULL,
            reactor_jid  TEXT,
            reactor_name TEXT,
            emoji        TEXT NOT NULL,
            timestamp    INTEGER,
            source       TEXT NOT NULL DEFAULT 'baileys'
        );
        CREATE INDEX IF NOT EXISTS idx_wa_react ON whatsapp_reactions(target_msg_id);

        CREATE TABLE IF NOT EXISTS whatsapp_imports (
            file_sha256     TEXT PRIMARY KEY,
            filename        TEXT,
            imported_at     INTEGER NOT NULL,
            message_count   INTEGER NOT NULL DEFAULT 0,
            duplicate_count INTEGER NOT NULL DEFAULT 0,
            imported_by_sub TEXT NOT NULL
        );
        """)
    # Patch existing records for any configured name aliases
    if _NAME_ALIASES:
        with _lock, _conn() as db:
            for old_lower, new_name in _NAME_ALIASES.items():
                db.execute(
                    "UPDATE whatsapp_messages SET sender_name=? "
                    "WHERE lower(sender_name)=? AND sender_name!=?",
                    (new_name, old_lower, new_name),
                )
                db.execute(
                    "UPDATE whatsapp_reactions SET reactor_name=? "
                    "WHERE lower(reactor_name)=? AND reactor_name!=?",
                    (new_name, old_lower, new_name),
                )
            db.commit()
    logger.info("whatsapp_analytics: DB ready at %s", _DB_PATH)


# ── Parser ─────────────────────────────────────────────────────────────────────

# Standard Android: "DD/MM/YY, HH:MM - Sender: msg" (also MM/DD variants)
_PAT1 = re.compile(
    r'^(\d{1,2}[/\-]\d{1,2}[/\-]\d{2,4})'
    r',?\s+'
    r'(\d{1,2}:\d{2}(?::\d{2})?(?:\s*[AaPp]\.?[Mm]\.?)?)'
    r'\s*[-–—]\s+'
    r'([^:]+?):\s(.*)',
    re.DOTALL,
)
# iOS bracketed: "[DD/MM/YY, HH:MM:SS AM] Sender: msg"
_PAT2 = re.compile(
    r'^\[(\d{1,2}[/\-]\d{1,2}[/\-]\d{2,4})'
    r',?\s+'
    r'(\d{1,2}:\d{2}(?::\d{2})?(?:\s*[AaPp]\.?[Mm]\.?)?)'
    r'\]\s+'
    r'([^:]+?):\s(.*)',
    re.DOTALL,
)
# Patterns that are timestamp lines but NOT user messages (system notifications)
_SYS1 = re.compile(r'^\d{1,2}[/\-]\d{1,2}[/\-]\d{2,4},?\s+\d{1,2}:\d{2}')
_SYS2 = re.compile(r'^\[\d{1,2}[/\-]\d{1,2}[/\-]\d{2,4},?\s+\d{1,2}:\d{2}')

_MEDIA_RE = re.compile(
    r'<[Mm]edia omitted>|image omitted|video omitted|audio omitted|'
    r'sticker omitted|document omitted|GIF omitted|'
    r'<[Mm]ultimedia omitido>|imagen omitid',
    re.IGNORECASE,
)


def _parse_datetime(date_str: str, time_str: str) -> Optional[int]:
    date_str = date_str.strip()
    time_str = time_str.strip().upper().replace(".", "")
    ampm = ""
    if "AM" in time_str:
        ampm = "AM"
        time_str = time_str.replace("AM", "").strip()
    elif "PM" in time_str:
        ampm = "PM"
        time_str = time_str.replace("PM", "").strip()

    parts = time_str.split(":")
    if len(parts) < 2:
        return None
    try:
        hour, minute = int(parts[0]), int(parts[1])
        second = int(parts[2]) if len(parts) > 2 else 0
    except ValueError:
        return None
    if ampm == "PM" and hour != 12:
        hour += 12
    elif ampm == "AM" and hour == 12:
        hour = 0

    sep = "/" if "/" in date_str else "-"
    dp = date_str.split(sep)
    if len(dp) != 3:
        return None
    try:
        d1, d2, yr = int(dp[0]), int(dp[1]), dp[2]
    except ValueError:
        return None
    if len(yr) == 2:
        yr = "20" + yr
    try:
        year = int(yr)
    except ValueError:
        return None

    for day, month in [(d1, d2), (d2, d1)]:
        if 1 <= month <= 12 and 1 <= day <= 31:
            try:
                dt = datetime(year, month, day, hour, minute, second, tzinfo=_TZ)
                return int(dt.timestamp())
            except ValueError:
                continue
    return None


def parse_txt(text: str) -> list[dict]:
    """Parse a WhatsApp .txt export. Returns list of message dicts."""
    text = text.lstrip("﻿").replace("‎", "").replace("‏", "")
    lines = text.splitlines()
    messages: list[dict] = []
    current: dict | None = None

    for line in lines:
        m = _PAT1.match(line) or _PAT2.match(line)
        if m:
            if current:
                messages.append(current)
            date_s, time_s = m.group(1), m.group(2)
            sender = m.group(3).strip()
            body = (m.group(4) or "").strip()
            ts = _parse_datetime(date_s, time_s)
            if ts is None:
                current = None
                continue
            is_media = bool(_MEDIA_RE.search(body)) if body else False
            bl = body.lower()
            current = {
                "sender_name": _resolve_name(sender),
                "timestamp": ts,
                "text": None if is_media else (body or None),
                "is_media_omitted": is_media,
                "has_photo": 1 if is_media and ("image" in bl or "photo" in bl) else 0,
                "has_video": 1 if is_media and "video" in bl else 0,
                "has_audio": 1 if is_media and ("audio" in bl or "voice" in bl or "ptt" in bl) else 0,
                "has_document": 1 if is_media and "document" in bl else 0,
                "message_type": "media" if is_media else "text",
            }
        elif current is not None:
            # System notification (timestamp present but no colon-sender pattern)
            if _SYS1.match(line) or _SYS2.match(line):
                messages.append(current)
                current = None
            elif line.strip():
                # Multiline continuation
                current["text"] = ((current.get("text") or "") + "\n" + line).strip()

    if current:
        messages.append(current)
    return messages


# ── Import ─────────────────────────────────────────────────────────────────────

def _msg_id(sender: str, ts: int, text: str, batch: str) -> str:
    raw = f"{batch}|{sender}|{ts}|{(text or '')[:120]}"
    return hashlib.sha256(raw.encode()).hexdigest()


def import_messages(content: bytes, filename: str, group_jid: str, imported_by_sub: str) -> dict:
    """Parse and ingest a WhatsApp export (bytes). Returns import summary."""
    file_sha = hashlib.sha256(content).hexdigest()

    with _lock, _conn() as db:
        existing = db.execute(
            "SELECT message_count, duplicate_count FROM whatsapp_imports WHERE file_sha256=?",
            (file_sha,),
        ).fetchone()
    if existing:
        return {
            "status": "already_imported",
            "file_sha256": file_sha,
            "message_count": existing["message_count"],
            "duplicate_count": existing["duplicate_count"],
        }

    if filename.lower().endswith(".zip"):
        txt_content = _extract_txt_from_zip(content)
    else:
        txt_content = content.decode("utf-8", errors="replace")

    messages = parse_txt(txt_content)
    if not messages:
        return {"status": "no_messages", "file_sha256": file_sha, "message_count": 0}

    now = int(time.time())
    inserted = 0
    duplicates = 0
    batch_tag = file_sha[:16]

    with _lock, _conn() as db:
        for msg in messages:
            mid = _msg_id(msg["sender_name"], msg["timestamp"], msg.get("text") or "", batch_tag)
            cur = db.execute(
                "INSERT OR IGNORE INTO whatsapp_messages "
                "(id, message_id, group_jid, sender_jid, sender_name, timestamp, text, "
                "message_type, has_photo, has_video, has_audio, has_document, is_media_omitted, "
                "reply_to, from_me, source, import_batch, created_at) VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)",
                (mid, None, group_jid or None, None,
                 msg["sender_name"], msg["timestamp"],
                 msg.get("text"), msg.get("message_type", "text"),
                 msg.get("has_photo", 0), msg.get("has_video", 0),
                 msg.get("has_audio", 0), msg.get("has_document", 0),
                 1 if msg.get("is_media_omitted") else 0,
                 None, 0, "historical_export", batch_tag, now),
            )
            if cur.rowcount:
                inserted += 1
            else:
                duplicates += 1
        db.execute(
            "INSERT INTO whatsapp_imports (file_sha256, filename, imported_at, message_count, duplicate_count, imported_by_sub) "
            "VALUES (?,?,?,?,?,?)",
            (file_sha, filename, now, inserted, duplicates, imported_by_sub),
        )
        db.commit()

    logger.info("wa_import: file=%s inserted=%d dupes=%d sha=%s", filename, inserted, duplicates, file_sha[:12])
    return {
        "status": "imported",
        "file_sha256": file_sha,
        "message_count": inserted,
        "duplicate_count": duplicates,
        "total_parsed": len(messages),
    }


def _extract_txt_from_zip(content: bytes) -> str:
    uncompressed = 0
    parts: list[str] = []
    try:
        with zipfile.ZipFile(io.BytesIO(content)) as zf:
            names = [n for n in zf.namelist() if n.lower().endswith(".txt") and ".." not in n and not n.startswith("/")]
            names = names[:MAX_ZIP_FILES]
            for name in names:
                info = zf.getinfo(name)
                uncompressed += info.file_size
                if uncompressed > MAX_ZIP_UNCOMPRESSED:
                    raise ValueError("ZIP uncompressed content exceeds 150 MB limit")
                parts.append(zf.read(name).decode("utf-8", errors="replace"))
    except zipfile.BadZipFile as exc:
        raise ValueError("Not a valid ZIP file") from exc
    return "\n".join(parts)


# ── Baileys live ingest ────────────────────────────────────────────────────────

def ingest_baileys_message(msg: dict) -> bool:
    """Store a single Baileys message. Returns True if newly inserted."""
    message_id = msg.get("message_id") or msg.get("id") or ""
    sender_jid = msg.get("sender_jid") or msg.get("from") or ""
    sender_name = _resolve_name(
        msg.get("sender_name") or msg.get("pushName") or
        (sender_jid.split("@")[0] if sender_jid else "") or "Unknown"
    )
    group_jid = msg.get("group_jid") or msg.get("groupJid") or ""
    ts = msg.get("timestamp") or msg.get("messageTimestamp") or int(time.time())
    ts = int(ts)
    if ts > 9_999_999_999:  # milliseconds
        ts //= 1000
    text = msg.get("text") or msg.get("body") or None
    msg_type = msg.get("message_type") or msg.get("type") or "text"
    has_photo = int(bool(msg.get("has_photo") or msg_type == "image"))
    has_video = int(bool(msg.get("has_video") or msg_type == "video"))
    has_audio = int(bool(msg.get("has_audio") or msg_type in ("audio", "ptt")))
    has_document = int(bool(msg.get("has_document") or msg_type == "document"))
    reply_to = msg.get("reply_to") or msg.get("quotedMessageId") or None
    from_me = int(bool(msg.get("from_me") or msg.get("fromMe")))

    if message_id:
        row_id = hashlib.sha256(f"baileys|{message_id}".encode()).hexdigest()
    else:
        row_id = _msg_id(sender_name, ts, text or "", "baileys")

    now = int(time.time())
    with _lock, _conn() as db:
        if message_id:
            db.execute(
                "DELETE FROM whatsapp_messages WHERE message_id=? AND source='historical_export'",
                (message_id,),
            )
        cur = db.execute(
            "INSERT OR IGNORE INTO whatsapp_messages "
            "(id, message_id, group_jid, sender_jid, sender_name, timestamp, text, "
            "message_type, has_photo, has_video, has_audio, has_document, is_media_omitted, "
            "reply_to, from_me, source, import_batch, created_at) VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)",
            (row_id, message_id or None, group_jid or None, sender_jid or None,
             sender_name, ts, text, msg_type,
             has_photo, has_video, has_audio, has_document, 0,
             reply_to, from_me, "baileys", None, now),
        )
        db.commit()
        return cur.rowcount > 0


def ingest_baileys_reaction(reaction: dict) -> bool:
    """Store a reaction from Baileys."""
    target = reaction.get("target_msg_id") or reaction.get("id") or ""
    reactor_jid = reaction.get("reactor_jid") or reaction.get("from") or ""
    reactor_name = _resolve_name(reaction.get("reactor_name") or reactor_jid.split("@")[0] or "")
    emoji = reaction.get("emoji") or reaction.get("text") or ""
    if not emoji or not target:
        return False
    ts = int(reaction.get("timestamp") or time.time())
    rid = hashlib.sha256(f"react|{target}|{reactor_jid}|{emoji}".encode()).hexdigest()
    with _lock, _conn() as db:
        cur = db.execute(
            "INSERT OR IGNORE INTO whatsapp_reactions (id, target_msg_id, reactor_jid, reactor_name, emoji, timestamp, source) VALUES (?,?,?,?,?,?,?)",
            (rid, target, reactor_jid, reactor_name, emoji, ts, "baileys"),
        )
        db.commit()
        return cur.rowcount > 0


# ── Shared helpers ─────────────────────────────────────────────────────────────

def _ts_bounds(range_str: str, start: str = "", end: str = "") -> tuple[int | None, int | None]:
    now = datetime.now(_TZ)
    if range_str == "this_month":
        s = datetime(now.year, now.month, 1, tzinfo=_TZ)
        return int(s.timestamp()), None
    if range_str == "prev_month":
        m = now.month - 1 or 12
        y = now.year if now.month > 1 else now.year - 1
        s = datetime(y, m, 1, tzinfo=_TZ)
        _, days = calendar.monthrange(y, m)
        e = datetime(y, m, days, 23, 59, 59, tzinfo=_TZ)
        return int(s.timestamp()), int(e.timestamp())
    if range_str == "this_year":
        s = datetime(now.year, 1, 1, tzinfo=_TZ)
        return int(s.timestamp()), None
    if range_str == "custom" and start and end:
        try:
            s_dt = datetime.fromisoformat(start).replace(tzinfo=_TZ)
            e_dt = datetime.fromisoformat(end).replace(hour=23, minute=59, second=59, tzinfo=_TZ)
            return int(s_dt.timestamp()), int(e_dt.timestamp())
        except Exception:
            pass
    return None, None


def _where_ts(s: int | None, e: int | None) -> tuple[str, list]:
    parts, p = [], []
    if s:
        parts.append("m.timestamp >= ?"); p.append(s)
    if e:
        parts.append("m.timestamp <= ?"); p.append(e)
    return (" AND " + " AND ".join(parts)) if parts else "", p


def _q(sql: str, params=()) -> list[dict]:
    with _conn() as db:
        return [dict(r) for r in db.execute(sql, params).fetchall()]


# ── Stats ──────────────────────────────────────────────────────────────────────

def stats(range_str: str = "all_time", start: str = "", end: str = "") -> dict:
    s, e = _ts_bounds(range_str, start, end)
    w, p = _where_ts(s, e)
    total   = _q(f"SELECT COUNT(*) n FROM whatsapp_messages m WHERE 1=1{w}", p)[0]["n"]
    members = _q(f"SELECT COUNT(DISTINCT sender_name) n FROM whatsapp_messages m WHERE 1=1{w}", p)[0]["n"]
    videos  = _q(f"SELECT COUNT(*) n FROM whatsapp_messages m WHERE has_video=1{w}", p)[0]["n"]
    photos  = _q(f"SELECT COUNT(*) n FROM whatsapp_messages m WHERE has_photo=1{w}", p)[0]["n"]
    span    = _q(f"SELECT MIN(timestamp) mn, MAX(timestamp) mx FROM whatsapp_messages m WHERE 1=1{w}", p)[0]
    days = max(1, int((span["mx"] - span["mn"]) / 86400)) if span["mn"] and span["mx"] else 0
    per_member = _q(f"SELECT sender_name, COUNT(*) cnt FROM whatsapp_messages m WHERE 1=1{w} GROUP BY sender_name ORDER BY cnt DESC", p)
    return {
        "total_messages": total, "total_members": members,
        "total_videos": videos, "total_photos": photos,
        "conversation_days": days,
        "first_ts": span["mn"], "last_ts": span["mx"],
        "member_message_counts": per_member,
    }


# ── Activity ───────────────────────────────────────────────────────────────────

def activity(range_str: str = "all_time", start: str = "", end: str = "") -> dict:
    s, e = _ts_bounds(range_str, start, end)
    w, p = _where_ts(s, e)
    rows = _q(f"SELECT timestamp, sender_name FROM whatsapp_messages m WHERE 1=1{w} ORDER BY timestamp", p)

    hour_c: Counter = Counter()
    dow_c: Counter = Counter()
    daily: dict = defaultdict(int)
    monthly: dict = defaultdict(int)
    member_monthly: dict = defaultdict(lambda: defaultdict(int))

    for r in rows:
        dt = datetime.fromtimestamp(r["timestamp"], tz=_TZ)
        hour_c[dt.hour] += 1
        dow_c[dt.weekday()] += 1
        daily[dt.strftime("%Y-%m-%d")] += 1
        monthly[dt.strftime("%Y-%m")] += 1
        member_monthly[r["sender_name"]][dt.strftime("%Y-%m")] += 1

    top_days = sorted(daily.items(), key=lambda x: x[1], reverse=True)[:5]

    return {
        "by_hour": [{"hour": h, "count": hour_c.get(h, 0)} for h in range(24)],
        "by_dow": [{"dow": d, "label": ["Mon","Tue","Wed","Thu","Fri","Sat","Sun"][d], "count": dow_c.get(d, 0)} for d in range(7)],
        "daily": [{"date": k, "count": v} for k, v in sorted(daily.items())],
        "monthly": [{"month": k, "count": v} for k, v in sorted(monthly.items())],
        "top_days": [{"date": k, "count": v} for k, v in top_days],
        "member_monthly": {n: dict(m) for n, m in member_monthly.items()},
    }


# ── Heatmap ────────────────────────────────────────────────────────────────────

def heatmap(range_str: str = "all_time", start: str = "", end: str = "") -> dict:
    s, e = _ts_bounds(range_str, start, end)
    w, p = _where_ts(s, e)
    rows = _q(f"SELECT timestamp FROM whatsapp_messages m WHERE 1=1{w}", p)
    cells: dict = defaultdict(int)
    for r in rows:
        dt = datetime.fromtimestamp(r["timestamp"], tz=_TZ)
        cells[(dt.weekday(), dt.hour)] += 1
    mx = max(cells.values()) if cells else 1
    return {
        "cells": [{"dow": d, "hour": h, "count": c} for (d, h), c in cells.items()],
        "max_count": mx,
    }


# ── Words ──────────────────────────────────────────────────────────────────────

_SW = frozenset({
    "the","and","you","that","was","for","are","with","his","her","they","this",
    "have","from","one","had","not","but","what","all","were","when","there","been",
    "your","which","their","said","she","each","how","will","other","about","out",
    "many","then","them","these","some","into","who","its","like","just","him",
    "did","get","yes","hey","okay","haha","lol","yeah","ok","oh","ah","omg","ugh",
    "omitted","media","image","video","audio","document","sticker","gif",
    "https","http","www","com","can","well","also","even","only","know","think",
    "time","want","good","going","really","got","now","still","way","make","would",
    "could","should","more","much","very","too","over","after","before","though",
    "actually","something","anything","everything","nothing","someone","anyone",
    "sorry","thanks","thank","please","come","send","sent","gonna","wanna","dont",
    "cant","doesnt","isnt","arent","wouldnt","couldnt","shouldnt","must","has",
    "haha","hahaha","lmao","bruh","bro","man","dude","rn","ngl","imo","tbh","idk",
    "nvm","btw","smh","ong","literally","basically","totally","kinda","sorta","ugh",
    "ahh","hmm","umm","nah","yep","yea","nope","mhm","ohh","wow","omg","wtf",
    "that","with","this","they","from","been","have","what","when","there","which",
    "their","will","than","then","them","some","into","over","after","before","were",
    "being","about","also","only","even","like","just","still","very","too","out",
    "all","now","any","our","use","can","may","let","per","see","way","get","its",
    "here","both","same","long","back","well","made","take","give","most","yes",
    "how","why","who","what","where","when","thus","else","ever","yet","for","nor",
    "are","was","had","did","has","him","her","his","she","they","them","their",
    "its","who","whom","each","few","more","most","other","some","such","into",
    "off","own","two","new","one","all","but","not","own","out","far",
})


def _tokenize(text: str) -> list[str]:
    if not text:
        return []
    text = re.sub(r'https?://\S+|www\.\S+', ' ', text)
    text = re.sub(r'[\U00010000-\U0010FFFF\U00002000-\U0001FFFF]', ' ', text, flags=re.UNICODE)
    text = re.sub(r'[^a-zA-ZÀ-ÿ\s]', ' ', text)
    return [w for w in text.lower().split() if len(w) >= 3 and w not in _SW]


def words(range_str: str = "all_time", start: str = "", end: str = "", limit: int = 50) -> dict:
    s, e = _ts_bounds(range_str, start, end)
    w, p = _where_ts(s, e)
    rows = _q(f"SELECT sender_name, text FROM whatsapp_messages m WHERE text IS NOT NULL AND is_media_omitted=0{w}", p)
    all_w: Counter = Counter()
    per: dict = defaultdict(Counter)
    for r in rows:
        toks = _tokenize(r["text"] or "")
        all_w.update(toks)
        per[r["sender_name"]].update(toks)
    return {
        "top_words": [{"word": w, "count": c} for w, c in all_w.most_common(limit)],
        "member_top_words": {n: [{"word": ww, "count": cc} for ww, cc in ctr.most_common(15)] for n, ctr in per.items()},
    }


# ── Emoji ──────────────────────────────────────────────────────────────────────

def _is_emoji_cp(cp: int) -> bool:
    return (
        0x1F300 <= cp <= 0x1F9FF or 0x1FA00 <= cp <= 0x1FAFF or
        0x2600  <= cp <= 0x27BF  or 0x2B00  <= cp <= 0x2BFF  or
        0xFE00  <= cp <= 0xFE0F  or
        cp in (0x231A, 0x231B, 0x23E9, 0x23F3, 0x25B6, 0x25C0, 0x25AA, 0x25AB,
               0x25FB, 0x25FC, 0x25FD, 0x25FE, 0x2614, 0x2615, 0x2693, 0x26A1,
               0x26AA, 0x26AB, 0x26BD, 0x26BE, 0x26C4, 0x26C5, 0x26D4, 0x26EA,
               0x26F2, 0x26F3, 0x26F5, 0x26FA, 0x26FD, 0x2702, 0x2705, 0x2734,
               0x2747, 0x274C, 0x274E, 0x2753, 0x2757, 0x2795, 0x2796, 0x2797,
               0x27A1, 0x27B0, 0x27BF, 0x2934, 0x2935, 0x2B1B, 0x2B1C, 0x2B50,
               0x2B55, 0x3030, 0x303D, 0x3297, 0x3299)
    )


def _extract_emoji(text: str) -> list[str]:
    result = []
    i = 0
    while i < len(text):
        cp = ord(text[i])
        if _is_emoji_cp(cp):
            seq = text[i]
            i += 1
            while i < len(text):
                ncp = ord(text[i])
                if ncp in (0x200D, 0xFE0F, 0x20E3) or 0x1F3FB <= ncp <= 0x1F3FF or 0x1F1E0 <= ncp <= 0x1F1FF:
                    seq += text[i]
                    i += 1
                else:
                    break
            result.append(seq)
        else:
            i += 1
    return result


def emojis(range_str: str = "all_time", start: str = "", end: str = "") -> dict:
    s, e = _ts_bounds(range_str, start, end)
    w, p = _where_ts(s, e)
    rows = _q(f"SELECT sender_name, text FROM whatsapp_messages m WHERE text IS NOT NULL{w}", p)
    all_e: Counter = Counter()
    per: dict = defaultdict(Counter)
    for r in rows:
        found = _extract_emoji(r["text"] or "")
        all_e.update(found)
        per[r["sender_name"]].update(found)
    total = sum(all_e.values())
    return {
        "top_emoji": [{"emoji": e, "count": c, "pct": round(c / total * 100, 1) if total else 0} for e, c in all_e.most_common(20)],
        "total_emoji": total,
        "member_top_emoji": {n: [{"emoji": e, "count": c} for e, c in ctr.most_common(8)] for n, ctr in per.items()},
    }


# ── Response Times ─────────────────────────────────────────────────────────────

def response_times(range_str: str = "all_time", start: str = "", end: str = "") -> dict:
    s, e = _ts_bounds(range_str, start, end)
    w, p = _where_ts(s, e)
    rows = _q(f"SELECT sender_name, timestamp FROM whatsapp_messages m WHERE 1=1{w} ORDER BY timestamp", p)
    if len(rows) < 10:
        return {"member_avg_minutes": [], "distribution": [], "fastest_responder": None}

    events: list[dict] = []
    for i in range(1, len(rows)):
        prev, curr = rows[i - 1], rows[i]
        if prev["sender_name"] != curr["sender_name"]:
            diff = (curr["timestamp"] - prev["timestamp"]) / 60.0
            if 0 < diff <= 60:
                events.append({"responder": curr["sender_name"], "minutes": diff})

    if not events:
        return {"member_avg_minutes": [], "distribution": [], "fastest_responder": None}

    per: dict = defaultdict(list)
    for ev in events:
        per[ev["responder"]].append(ev["minutes"])

    member_avg = sorted(
        [{"name": n, "avg_minutes": round(sum(t) / len(t), 1), "count": len(t)} for n, t in per.items()],
        key=lambda x: x["avg_minutes"],
    )
    all_t = [ev["minutes"] for ev in events]
    dist = [{"label": f"{lo}-{hi}m", "count": sum(1 for t in all_t if lo <= t < hi)}
            for lo, hi in [(0, 5), (5, 10), (10, 20), (20, 30), (30, 60)]]

    return {
        "member_avg_minutes": member_avg,
        "distribution": dist,
        "fastest_responder": member_avg[0]["name"] if member_avg else None,
        "event_count": len(events),
    }


# ── Members ────────────────────────────────────────────────────────────────────

def members(range_str: str = "all_time", start: str = "", end: str = "") -> dict:
    s, e = _ts_bounds(range_str, start, end)
    w, p = _where_ts(s, e)
    rows = _q(
        f"SELECT sender_name, COUNT(*) msgs, SUM(has_photo) photos, SUM(has_video) videos, "
        f"SUM(has_audio) audios, SUM(is_media_omitted) media_omit, MIN(timestamp) first_ts, MAX(timestamp) last_ts "
        f"FROM whatsapp_messages m WHERE 1=1{w} GROUP BY sender_name ORDER BY msgs DESC", p
    )
    text_rows = _q(f"SELECT sender_name, text FROM whatsapp_messages m WHERE text IS NOT NULL AND is_media_omitted=0{w}", p)
    wc: dict = defaultdict(int)
    cc: dict = defaultdict(int)
    for r in text_rows:
        t = r["text"] or ""
        wc[r["sender_name"]] += len(t.split())
        cc[r["sender_name"]] += len(t)
    return {"members": [{
        "name": r["sender_name"], "messages": r["msgs"] or 0,
        "photos": r["photos"] or 0, "videos": r["videos"] or 0,
        "audios": r["audios"] or 0, "media_omitted": r["media_omit"] or 0,
        "total_words": wc.get(r["sender_name"], 0),
        "total_chars": cc.get(r["sender_name"], 0),
        "avg_words_per_msg": round(wc.get(r["sender_name"], 0) / (r["msgs"] or 1), 1),
        "first_ts": r["first_ts"], "last_ts": r["last_ts"],
    } for r in rows]}


# ── Awards ─────────────────────────────────────────────────────────────────────

def awards(range_str: str = "all_time", start: str = "", end: str = "") -> dict:
    s, e = _ts_bounds(range_str, start, end)
    w, p = _where_ts(s, e)

    def _top(extra: str = "", ep: list | None = None):
        ep = ep or []
        r = _q(f"SELECT sender_name, COUNT(*) cnt FROM whatsapp_messages m WHERE 1=1{w}{extra} GROUP BY sender_name ORDER BY cnt DESC LIMIT 1", p + ep)
        return r[0] if r else None

    def _emoji_top(ch: str):
        r = _q(f"SELECT sender_name, COUNT(*) cnt FROM whatsapp_messages m WHERE text LIKE ?{w} GROUP BY sender_name ORDER BY cnt DESC LIMIT 1", [f"%{ch}%"] + p)
        return r[0] if r else None

    ts_rows = _q(f"SELECT sender_name, timestamp FROM whatsapp_messages m WHERE 1=1{w}", p)
    night_c: Counter = Counter()
    early_c: Counter = Counter()
    hour_c: Counter = Counter()
    day_c: Counter = Counter()
    all_e: Counter = Counter()

    text_rows = _q(f"SELECT text FROM whatsapp_messages m WHERE text IS NOT NULL{w}", p)
    for r in text_rows:
        all_e.update(_extract_emoji(r["text"] or ""))

    for r in ts_rows:
        dt = datetime.fromtimestamp(r["timestamp"], tz=_TZ)
        h = dt.hour
        if 0 <= h <= 3:
            night_c[r["sender_name"]] += 1
        elif 5 <= h <= 8:
            early_c[r["sender_name"]] += 1
        hour_c[h] += 1
        day_c[dt.strftime("%Y-%m-%d")] += 1

    # Longest streak
    all_days = sorted({datetime.fromtimestamp(r["timestamp"], tz=_TZ).date() for r in ts_rows})
    streak = best_streak = (1 if all_days else 0)
    for i in range(1, len(all_days)):
        if (all_days[i] - all_days[i - 1]).days == 1:
            streak += 1
            best_streak = max(best_streak, streak)
        else:
            streak = 1

    # Ghost of the month (always current month)
    gms, gme = _ts_bounds("this_month")
    gw, gp = _where_ts(gms, gme)
    ghost_row = _q(f"SELECT sender_name, COUNT(*) cnt FROM whatsapp_messages m WHERE 1=1{gw} GROUP BY sender_name ORDER BY cnt ASC LIMIT 1", gp)

    rt = response_times(range_str, start, end)
    react_rows = _q(
        f"SELECT m.sender_name, COUNT(r.id) cnt FROM whatsapp_messages m "
        f"JOIN whatsapp_reactions r ON r.target_msg_id=m.message_id "
        f"WHERE m.message_id IS NOT NULL{w} GROUP BY m.sender_name ORDER BY cnt DESC LIMIT 1", p
    )
    react_msg = _q(
        f"SELECT m.sender_name, m.text, m.timestamp, COUNT(r.id) cnt FROM whatsapp_messages m "
        f"JOIN whatsapp_reactions r ON r.target_msg_id=m.message_id "
        f"WHERE m.message_id IS NOT NULL{w} GROUP BY m.id ORDER BY cnt DESC LIMIT 1", p
    )

    def _n(row, key="sender_name"):
        return row[key] if row else None
    def _c(row):
        return row["cnt"] if row else None

    yapper = _top()
    video_king = _top(" AND has_video=1")
    photo_king = _top(" AND has_photo=1")
    skull = _emoji_top("💀")
    laugh = _emoji_top("😂")
    fire  = _emoji_top("🔥")

    return {
        "certified_yapper":    {"name": _n(yapper),     "count": _c(yapper)}    if yapper    else None,
        "night_owl":           {"name": night_c.most_common(1)[0][0], "count": night_c.most_common(1)[0][1]} if night_c else None,
        "early_bird":          {"name": early_c.most_common(1)[0][0], "count": early_c.most_common(1)[0][1]} if early_c else None,
        "video_king":          {"name": _n(video_king), "count": _c(video_king)} if video_king else None,
        "photo_king":          {"name": _n(photo_king), "count": _c(photo_king)} if photo_king else None,
        "most_reacted_person": {"name": react_rows[0]["sender_name"], "count": react_rows[0]["cnt"]} if react_rows else None,
        "most_reacted_message": react_msg[0] if react_msg else None,
        "most_used_emoji":     {"emoji": all_e.most_common(1)[0][0], "count": all_e.most_common(1)[0][1]} if all_e else None,
        "most_skull":          {"name": _n(skull),  "count": _c(skull)}  if skull else None,
        "most_laugh":          {"name": _n(laugh),  "count": _c(laugh)}  if laugh else None,
        "most_fire":           {"name": _n(fire),   "count": _c(fire)}   if fire  else None,
        "peak_hour":           {"hour": hour_c.most_common(1)[0][0], "count": hour_c.most_common(1)[0][1]} if hour_c else None,
        "biggest_day":         {"date": day_c.most_common(1)[0][0], "count": day_c.most_common(1)[0][1]} if day_c else None,
        "longest_streak_days": best_streak,
        "fastest_replier":     {"name": rt["member_avg_minutes"][0]["name"], "avg_minutes": rt["member_avg_minutes"][0]["avg_minutes"]} if rt.get("member_avg_minutes") else None,
        "ghost_of_month":      {"name": ghost_row[0]["sender_name"], "count": ghost_row[0]["cnt"]} if ghost_row else None,
    }


# ── Excel Export ───────────────────────────────────────────────────────────────

def export_xlsx(range_str: str = "all_time", start: str = "", end: str = "") -> bytes:
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill
        from openpyxl.utils import get_column_letter
    except ImportError as exc:
        raise RuntimeError("openpyxl not installed — add it to requirements.txt") from exc

    s, e = _ts_bounds(range_str, start, end)
    wq, p = _where_ts(s, e)
    wb = openpyxl.Workbook()

    hf = Font(color="FF2FD6", bold=True)
    hfill = PatternFill("solid", fgColor="120A26")

    def _sheet(ws, headers, rows):
        ws.append(headers)
        for cell in ws[1]:
            cell.font = hf
            cell.fill = hfill
        for row in rows:
            ws.append(list(row))
        for i in range(1, len(headers) + 1):
            ws.column_dimensions[get_column_letter(i)].width = 20

    ws = wb.active
    ws.title = "Messages"
    msgs = _q(f"SELECT sender_name, timestamp, text, message_type, has_photo, has_video, has_audio, source FROM whatsapp_messages m WHERE 1=1{wq} ORDER BY timestamp", p)
    _sheet(ws, ["Sender","Date","Message","Type","Photo","Video","Audio","Source"],
           [(r["sender_name"], datetime.fromtimestamp(r["timestamp"], tz=_TZ).strftime("%Y-%m-%d %H:%M"),
             (r["text"] or "")[:500], r["message_type"],
             r["has_photo"], r["has_video"], r["has_audio"], r["source"]) for r in msgs])

    ws2 = wb.create_sheet("Member Stats")
    md = members(range_str, start, end)["members"]
    _sheet(ws2, ["Name","Messages","Photos","Videos","Words","Avg Words/Msg"],
           [(m["name"], m["messages"], m["photos"], m["videos"], m["total_words"], m["avg_words_per_msg"]) for m in md])

    ws3 = wb.create_sheet("Activity by Hour")
    act = activity(range_str, start, end)
    _sheet(ws3, ["Hour","Messages"], [(r["hour"], r["count"]) for r in act["by_hour"]])

    ws4 = wb.create_sheet("Activity by Day")
    _sheet(ws4, ["Day","Messages"], [(r["label"], r["count"]) for r in act["by_dow"]])

    ws5 = wb.create_sheet("Monthly Activity")
    _sheet(ws5, ["Month","Messages"], [(r["month"], r["count"]) for r in act["monthly"]])

    ws6 = wb.create_sheet("Top Words")
    wd = words(range_str, start, end, limit=100)
    _sheet(ws6, ["Word","Count"], [(r["word"], r["count"]) for r in wd["top_words"]])

    ws7 = wb.create_sheet("Top Emoji")
    ed = emojis(range_str, start, end)
    _sheet(ws7, ["Emoji","Count","Percent"], [(r["emoji"], r["count"], r["pct"]) for r in ed["top_emoji"]])

    ws8 = wb.create_sheet("Response Times")
    rt = response_times(range_str, start, end)
    _sheet(ws8, ["Responder","Avg Min","Count"],
           [(r["name"], r["avg_minutes"], r["count"]) for r in rt.get("member_avg_minutes", [])])

    ws9 = wb.create_sheet("Response Distribution")
    _sheet(ws9, ["Range","Count"], [(r["label"], r["count"]) for r in rt.get("distribution", [])])

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
